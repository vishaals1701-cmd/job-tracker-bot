from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, CommandHandler, filters, ContextTypes
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

FILE_NAME = "jobs.xlsx"

VALID_STATUS = ["applied", "in progress", "rejected", "selected"]

# -------- CREATE FILE --------
if not os.path.exists(FILE_NAME):
    df = pd.DataFrame(columns=[
        "ID","User","Source","Company","Role","Applied_Date","Status"
    ])
    df.to_excel(FILE_NAME, index=False)

# -------- LOAD --------
def load_data():
    df = pd.read_excel(FILE_NAME)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    df = df[df["Company"].notna()]
    return df

# -------- SAVE + FORMAT --------
def save_data(df):
    df.to_excel(FILE_NAME, index=False)

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    # -------- HEADER COLOR --------
    header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill

    # -------- AUTO WIDTH (ALL COLUMNS) --------
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_len + 3

    # -------- STATUS COLORS --------
    yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    red = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    green = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        cell = row[6]
        if cell.value:
            status = str(cell.value).lower()
            if status == "applied":
                cell.fill = yellow
            elif status == "in progress":
                cell.fill = blue
            elif status == "rejected":
                cell.fill = red
            elif status == "selected":
                cell.fill = green

    # -------- CLEAR OLD SUMMARY --------
    for row in ws.iter_rows(min_row=1, max_row=50, min_col=10, max_col=11):
        for cell in row:
            cell.value = None

    # -------- SUMMARY DATA --------
    df["Status"] = df["Status"].str.lower()
    df["Source"] = df["Source"].str.lower()

    summary = [
        ("Total Applied", len(df)),
        ("SLA Applied", len(df[df["Source"]=="sla"])),
        ("Off-Campus Applied", len(df[df["Source"]=="off-campus"])),
        ("Selected (SLA)", len(df[(df["Status"]=="selected") & (df["Source"]=="sla")])),
        ("Selected (Off-Campus)", len(df[(df["Status"]=="selected") & (df["Source"]=="off-campus")])),
        ("In Progress (SLA)", len(df[(df["Status"]=="in progress") & (df["Source"]=="sla")])),
        ("In Progress (Off-Campus)", len(df[(df["Status"]=="in progress") & (df["Source"]=="off-campus")])),
        ("Rejected (SLA)", len(df[(df["Status"]=="rejected") & (df["Source"]=="sla")])),
        ("Rejected (Off-Campus)", len(df[(df["Status"]=="rejected") & (df["Source"]=="off-campus")])),
    ]

    start_col = 10
    start_row = 2

    # -------- SUMMARY STYLES --------
    summary_bg = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    # Title
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row, end_column=start_col+1)

    title = ws.cell(row=start_row, column=start_col)
    title.value = "JOB SUMMARY"
    title.fill = summary_bg
    title.border = border
    title.alignment = center

    # Rows
    for i, (label, value) in enumerate(summary):
        r = start_row + i + 1

        c1 = ws.cell(row=r, column=start_col)
        c1.value = label
        c1.fill = summary_bg
        c1.border = border
        c1.alignment = left

        c2 = ws.cell(row=r, column=start_col + 1)
        c2.value = value
        c2.fill = summary_bg
        c2.border = border
        c2.alignment = center

    # -------- FORCE AUTO WIDTH FOR J & K --------
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 15

    wb.save(FILE_NAME)

# -------- PARSE --------
def parse_message(text):
    data = {}
    for line in text.split("\n"):
        if "Source:" in line:
            data["Source"] = line.split(":")[1].strip()
        elif "Company:" in line:
            data["Company"] = line.split(":")[1].strip()
        elif "Role:" in line:
            data["Role"] = line.split(":")[1].strip()
        elif "Date:" in line:
            data["Applied_Date"] = line.split(":")[1].strip()
        elif "Status:" in line:
            data["Status"] = line.split(":")[1].strip()
    return data

# -------- ADD --------
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user.username or "unknown"
    data = parse_message(update.message.text)

    df = load_data()
    new_id = 1 if df.empty else int(df["ID"].max()) + 1

    df = pd.concat([df, pd.DataFrame([{
        "ID": new_id,
        "User": user,
        "Source": data.get("Source"),
        "Company": data.get("Company"),
        "Role": data.get("Role"),
        "Applied_Date": data.get("Applied_Date"),
        "Status": data.get("Status")
    }])], ignore_index=True)

    save_data(df)
    await update.message.reply_text(f"✅ Job saved (ID: {new_id})")

# -------- VIEW --------
async def view_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = load_data()

    msg = "📋 JOB LIST\n\n"
    for _, row in df.iterrows():
        msg += f"{int(row['ID'])} | {row['Company']} | {row['Status']}\n"

    await update.message.reply_text(msg)

# -------- UPDATE --------
async def update_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    app_id = int(context.args[0])
    status = " ".join(context.args[1:]).lower()

    df = load_data()
    df.loc[df["ID"] == app_id, "Status"] = status.title()

    save_data(df)
    await update.message.reply_text("🔄 Updated")

# -------- DELETE --------
async def delete_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    app_id = int(context.args[0])

    df = load_data()
    df = df[df["ID"] != app_id]
    df = df.reset_index(drop=True)
    df["ID"] = df.index + 1

    save_data(df)
    await update.message.reply_text("🗑 Deleted")

# -------- SUMMARY --------
async def summary_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = load_data()

    df["Status"] = df["Status"].str.lower()
    df["Source"] = df["Source"].str.lower()

    msg = f"""📊 JOB SUMMARY

Total Applied: {len(df)}

SLA Applied: {len(df[df['Source']=='sla'])}
Off-Campus Applied: {len(df[df['Source']=='off-campus'])}

Selected (SLA): {len(df[(df['Status']=='selected') & (df['Source']=='sla')])}
Selected (Off-Campus): {len(df[(df['Status']=='selected') & (df['Source']=='off-campus')])}

In Progress (SLA): {len(df[(df['Status']=='in progress') & (df['Source']=='sla')])}
In Progress (Off-Campus): {len(df[(df['Status']=='in progress') & (df['Source']=='off-campus')])}

Rejected (SLA): {len(df[(df['Status']=='rejected') & (df['Source']=='sla')])}
Rejected (Off-Campus): {len(df[(df['Status']=='rejected') & (df['Source']=='off-campus')])}
"""

    await update.message.reply_text(msg)

# -------- MAIN --------
app = ApplicationBuilder().token("8776312749:AAFU3e0YJL8_wHXsDQycHom8DQxdMM9kjBY").build()

app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
app.add_handler(CommandHandler("view", view_cmd))
app.add_handler(CommandHandler("update", update_cmd))
app.add_handler(CommandHandler("delete", delete_cmd))
app.add_handler(CommandHandler("summary", summary_cmd))

print("🚀 Bot running...")
app.run_polling()



